"""
apps/ventas/nota_pedido_doc.py
Nota de Pedido como documento imprimible: PDF (reportlab) y Excel (openpyxl).

El diseño replica la nota que el negocio ya venía usando (ver
`docs/Ejemplos Nota de Pedido/`): logo arriba, título y cajas de fecha a la
derecha, datos del cliente sobre líneas, tabla CANTIDAD / PRODUCTO / PRECIO /
TOTAL que se completa con renglones en blanco hasta el pie, caja de total en
beige y franja beige de contacto con el monograma sangrando por el borde.

Los dos formatos NO son el mismo documento con otra extensión:

  · PDF   — documento de presentación. Respeta la diagramación del ejemplo
            (cuatro columnas, la cantidad y su unidad en una sola celda a
            dos líneas). Es lo que se imprime o se le manda al cliente.
  · Excel — documento de trabajo. Separa la unidad en su propia columna y
            deja CANTIDAD y PRECIO como números con TOTAL por fórmula
            (=cantidad*precio) más filas en blanco al final, para que se
            pueda seguir cargando y recalcule solo.

La paleta sale del propio ejemplo:
    marca #B99C74   beige #EBDBC0   marrón #4D2610   gris #545454
El #B99C74 es el marrón claro del logo — el mismo que la interfaz web ya usa
como color de acento, así que pantalla y papel quedan en el mismo tono.
"""
import io
import os
from datetime import date, datetime
from decimal import Decimal

from django.conf import settings
from django.http import HttpResponse

# ── Paleta de la marca ────────────────────────────────────────────────────────
MARCA  = '#B99C74'
BEIGE  = '#EBDBC0'
MARRON = '#4D2610'
GRIS   = '#545454'
NEGRO  = '#000000'
LINEA  = '#8C8C8C'

# El documento que el negocio le entrega al cliente antes de cerrar la venta se
# llama "nota de presupuesto" — así está titulado el que venían usando. Una vez
# confirmado, el mismo pedido se imprime como "nota de pedido". Es el mismo
# diseño con otro encabezado, por eso va como parámetro y no como dos módulos.
TIPO_PRESUPUESTO = 'presupuesto'
TIPO_PEDIDO      = 'pedido'
TITULOS = {
    TIPO_PRESUPUESTO: 'NOTA DE PRESUPUESTO',
    TIPO_PEDIDO:      'NOTA DE PEDIDO',
}
TIPO_DEFECTO = TIPO_PRESUPUESTO

_ASSETS = os.path.join(os.path.dirname(__file__), 'assets')
LOGO       = os.path.join(_ASSETS, 'logo_oga_pora.png')   # marca completa
LOGO_MARCA = os.path.join(_ASSETS, 'logo_marca.png')      # monograma, para el pie


# ══════════════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════════════
def _gs(valor) -> str:
    """1234567 → '1.234.567' (separador de miles paraguayo)."""
    try:
        return f'{int(round(float(valor))):,}'.replace(',', '.')
    except (TypeError, ValueError):
        return '0'


def _cantidad(valor) -> str:
    """Muestra 17 en vez de 17,00 pero conserva los decimales de 24,99."""
    try:
        f = float(valor)
    except (TypeError, ValueError):
        return '0'
    return f'{f:.0f}' if abs(f - round(f)) < 0.005 else f'{f:.2f}'


def _empresa() -> dict:
    """
    Datos del pie de la nota. Son los de CONTACTO_COMERCIAL, no los fiscales:
    el teléfono y la dirección que se le dan al cliente en el presupuesto no
    son los mismos que van en la factura.
    """
    fiscal   = getattr(settings, 'DATOS_FISCALES', {}) or {}
    contacto = getattr(settings, 'CONTACTO_COMERCIAL', {}) or {}
    return {
        'razon_social': fiscal.get('razon_social') or 'ÓGA PORÃ',
        'email':        contacto.get('email', ''),
        'direccion':    contacto.get('direccion', ''),
        'telefono':     contacto.get('telefono', ''),
    }


def _unidad_item(item) -> str:
    """
    Texto de unidad de un ítem: 'mts2 (17 cajas)' para porcelanato vendido
    por m², vacío para lo que se vende por pieza (ahí la cantidad se explica
    sola). Refleja cómo la nota de ejemplo escribe la columna CANTIDAD.
    """
    producto = item.variante.producto
    unidad   = producto.unidad_venta
    if unidad == producto.UNIDAD_PIEZA:
        return ''
    if unidad != producto.UNIDAD_M2:
        return dict(producto.UNIDADES).get(unidad, '').lower()

    m2_caja = item.variante.m2_por_caja_calculado
    if not m2_caja:
        return 'mts2'
    import math
    cajas = math.ceil(float(item.cantidad) / m2_caja)
    return f'mts2 ({cajas} caja{"s" if cajas != 1 else ""})'


# ══════════════════════════════════════════════════════════════════════
# Armado de los datos
# ══════════════════════════════════════════════════════════════════════
def datos_desde_pedido(pedido, tipo: str = TIPO_DEFECTO) -> dict:
    """
    Traduce una NotaPedido a la estructura plana que consumen los dos
    renderizadores. Mantenerlos desacoplados del ORM permite además generar
    una nota de prueba sin base de datos.

    `tipo` elige el encabezado: 'presupuesto' (lo que se le pasa al cliente
    para que decida) o 'pedido' (la venta ya confirmada).
    """
    items = []
    for item in pedido.items.select_related('variante__producto', 'variante__acabado').all():
        variante = item.variante
        partes   = [variante.producto.nombre]
        if variante.dimension_display:
            partes.append(variante.dimension_display)
        if variante.color:
            partes.append(variante.color)
        if getattr(variante, 'acabado', None):
            partes.append(variante.acabado.nombre)

        items.append({
            'cantidad':    Decimal(item.cantidad),
            'unidad':      _unidad_item(item),
            'descripcion': ' — '.join(p for p in partes if p),
            'precio':      Decimal(item.precio_unitario),
            'total':       Decimal(item.subtotal),
        })

    cliente = pedido.cliente
    return {
        'tipo':          tipo if tipo in TITULOS else TIPO_DEFECTO,
        'titulo':        TITULOS.get(tipo, TITULOS[TIPO_DEFECTO]),
        'numero':        pedido.numero,
        'fecha':         pedido.fecha_creacion.date() if pedido.fecha_creacion else date.today(),
        'cliente':       pedido.cliente_nombre or (cliente.razon_social if cliente else ''),
        'ruc':           pedido.cliente_ruc or (cliente.ruc if cliente else ''),
        'telefono':      pedido.cliente_telefono or (cliente.telefono if cliente else ''),
        'direccion':     cliente.direccion if cliente else '',
        'ciudad':        '',
        'vendedor':      pedido.vendedor.nombre_completo if pedido.vendedor else '',
        'observaciones': pedido.cliente_observaciones or '',
        'items':         items,
        'descuento':     Decimal(pedido.descuento or 0),
        'total':         Decimal(pedido.monto_a_cobrar or 0),
    }


# ══════════════════════════════════════════════════════════════════════
# PDF
# ══════════════════════════════════════════════════════════════════════
# Geometría en puntos sobre hoja A4, medida sobre la nota original que el
# negocio ya usaba (docs/Ejemplos Nota de Pedido/). No son valores libres:
# cambiarlos desalinea el documento respecto del que el cliente conoce.
IZQ, DER   = 51.8, 539.0
COLS       = [51.8, 121.2, 390.7, 449.1, 539.0]   # bordes de las 4 columnas
FOOTER_H   = 136.0
TABLA_TOP  = 576.0
ALTO_TH    = 30.0        # alto de la fila de encabezado de la tabla
ALTO_FILA  = 29.0        # alto de una fila de una sola línea
TOTAL_H    = 28.5
TABLA_MIN_Y = FOOTER_H + TOTAL_H + 28.0   # piso de la tabla en la última hoja


def _wrap(texto, ancho, fuente, tamanio):
    """Corta `texto` en líneas que entren en `ancho` puntos."""
    from reportlab.pdfbase.pdfmetrics import stringWidth

    palabras = str(texto).split()
    if not palabras:
        return ['']
    lineas, actual = [], palabras[0]
    for palabra in palabras[1:]:
        prueba = f'{actual} {palabra}'
        if stringWidth(prueba, fuente, tamanio) <= ancho:
            actual = prueba
        else:
            lineas.append(actual)
            actual = palabra
    lineas.append(actual)
    return lineas


def _lineas_cantidad(item):
    """
    Columna CANTIDAD a dos renglones cuando el ítem se vende por m²:
    '24,99 mts2' arriba y '(17 cajas)' abajo, como en la nota original.
    """
    cant_txt = _cantidad(item['cantidad'])
    unidad   = (item.get('unidad') or '').strip()
    if unidad.startswith('mts2'):
        resto = unidad[4:].strip()
        return [f'{cant_txt} mts2'] + ([resto] if resto else [])
    if unidad:
        return [f'{cant_txt} {unidad}']
    return [cant_txt]


def render_pdf(datos: dict) -> bytes:
    from reportlab.lib.colors import HexColor
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas as rl_canvas

    W, H = A4
    buffer = io.BytesIO()
    c = rl_canvas.Canvas(buffer, pagesize=A4)
    c.setTitle(f"{datos['titulo']} {datos.get('numero', '')}".strip())

    c_beige, c_marron = HexColor(BEIGE), HexColor(MARRON)
    c_gris, c_negro, c_linea = HexColor(GRIS), HexColor(NEGRO), HexColor(LINEA)
    empresa = _empresa()

    # ── Iconos del pie (vectoriales: Helvetica no trae pictogramas) ────
    def icono_sobre(x, y):
        c.setLineWidth(0.9)
        c.rect(x, y, 10, 7, stroke=1, fill=0)
        c.lines([(x, y + 7, x + 5, y + 3), (x + 5, y + 3, x + 10, y + 7)])

    def icono_pin(x, y):
        c.setLineWidth(0.9)
        c.circle(x + 5, y + 5, 3.6, stroke=1, fill=0)
        c.lines([(x + 2.6, y + 3.2, x + 5, y - 1), (x + 7.4, y + 3.2, x + 5, y - 1)])

    def icono_tel(x, y):
        c.setLineWidth(0.9)
        c.roundRect(x + 2, y - 0.5, 6.5, 9.5, 2, stroke=1, fill=0)

    # ── Franja de contacto (va en todas las hojas) ────────────────────
    def franja_pie():
        c.setFillColor(c_beige)
        c.rect(0, 0, W, FOOTER_H, stroke=0, fill=1)

        # Monograma como marca de agua, sangrando por el borde derecho.
        # El recorte lo mantiene dentro de la franja: sin él la parte alta
        # del dibujo se derrama sobre el blanco de la hoja.
        if os.path.exists(LOGO_MARCA):
            c.saveState()
            recorte = c.beginPath()
            recorte.rect(0, 0, W, FOOTER_H)
            c.clipPath(recorte, stroke=0, fill=0)
            alto = FOOTER_H + 34
            c.drawImage(LOGO_MARCA, W - 118, -14, width=alto * 317 / 372, height=alto,
                        mask='auto', preserveAspectRatio=True, anchor='sw')
            c.restoreState()

        c.setFillColor(c_marron)
        c.setStrokeColor(c_marron)
        c.setFont('Helvetica-Bold', 10.5)
        c.drawString(IZQ + 38, FOOTER_H - 40, 'CONTACTO:')

        c.setFont('Helvetica', 9.5)
        y = FOOTER_H - 72
        lineas = [(icono_sobre, empresa['email']),
                  (icono_pin,   empresa['direccion']),
                  (icono_tel,   empresa['telefono'])]
        for icono, texto in lineas:
            if not texto:
                continue
            icono(IZQ + 38, y - 1)
            for linea in _wrap(texto, 320, 'Helvetica', 9.5):
                c.drawString(IZQ + 56, y, linea)
                y -= 12.5
            y -= 5.5

    # ── Encabezado ────────────────────────────────────────────────────
    def encabezado(continuacion=False):
        if os.path.exists(LOGO):
            ancho_logo = 140.0
            alto_logo  = ancho_logo * 703 / 900
            c.drawImage(LOGO, 126, 813 - alto_logo, width=ancho_logo, height=alto_logo,
                        mask='auto', preserveAspectRatio=True, anchor='nw')

        # Cajas de fecha [FECHA][DD][MM][AA]
        f = datos.get('fecha') or date.today()
        celdas = [('FECHA', 379.0, 443.0),
                  (f'{f.day:02d}', 443.0, 472.0),
                  (f'{f.month:02d}', 472.0, 501.0),
                  (f'{f.year % 100:02d}', 501.0, 531.0)]
        c.setStrokeColor(c_linea)
        c.setLineWidth(0.8)
        for i, (texto, x0, x1) in enumerate(celdas):
            c.rect(x0, 743.5, x1 - x0, 27, stroke=1, fill=0)
            c.setFillColor(c_gris if i == 0 else c_negro)
            c.setFont('Helvetica' if i == 0 else 'Helvetica-Bold', 10.5 if i == 0 else 12)
            c.drawCentredString((x0 + x1) / 2, 752.5, texto)

        c.setFillColor(c_negro)
        c.setFont('Helvetica-Bold', 17)
        c.drawRightString(DER, 715, datos['titulo'])

        numero = datos.get('numero') or ''
        if numero:
            c.setFillColor(c_gris)
            c.setFont('Helvetica', 9)
            c.drawRightString(DER, 701, f'N° {numero}' + ('  (cont.)' if continuacion else ''))

        # Datos del cliente sobre líneas
        filas = [
            (636.0, [('Cliente:',   datos.get('cliente'),   IZQ, 97.0,  299.0),
                     ('Teléfono:',  datos.get('telefono'),  348.0, 395.0, DER)]),
            (607.4, [('Dirección:', datos.get('direccion'), IZQ, 97.0,  299.0),
                     ('RUC / CI:',  datos.get('ruc'),       348.0, 395.0, DER)]),
        ]
        for y, campos in filas:
            for etiqueta, valor, x_lbl, x_val, x_fin in campos:
                c.setFillColor(c_gris)
                c.setFont('Helvetica', 9.5)
                c.drawString(x_lbl, y, etiqueta)
                c.setStrokeColor(c_linea)
                c.setLineWidth(0.7)
                c.line(x_val, y - 4.5, x_fin, y - 4.5)
                if valor:
                    c.setFillColor(c_negro)
                    c.setFont('Helvetica', 11)
                    c.drawString(x_val + 6, y, str(valor)[:44])

    # ── Tabla ─────────────────────────────────────────────────────────
    ENCABEZADOS = ['CANTIDAD', 'PRODUCTO', 'PRECIO', 'TOTAL']
    ANCHO_DESC  = COLS[2] - COLS[1] - 16
    ANCHO_CANT  = COLS[1] - COLS[0] - 8

    def encabezado_tabla():
        c.setFillColor(c_negro)
        c.setFont('Helvetica-Bold', 10.5)
        for i, titulo in enumerate(ENCABEZADOS):
            c.drawCentredString((COLS[i] + COLS[i + 1]) / 2,
                                TABLA_TOP - ALTO_TH + 10.5, titulo)
        c.setStrokeColor(c_linea)
        c.setLineWidth(0.7)
        c.line(COLS[0], TABLA_TOP, COLS[-1], TABLA_TOP)
        c.line(COLS[0], TABLA_TOP - ALTO_TH, COLS[-1], TABLA_TOP - ALTO_TH)
        return TABLA_TOP - ALTO_TH

    def cerrar_tabla(y_bot):
        """Verticales y borde inferior del bloque de tabla de esta hoja."""
        c.setStrokeColor(c_linea)
        c.setLineWidth(0.7)
        for x in COLS:
            c.line(x, TABLA_TOP, x, y_bot)
        c.line(COLS[0], y_bot, COLS[-1], y_bot)

    def dibujar_fila(item, y_top, alto):
        lineas_cant = [l for linea in _lineas_cantidad(item)
                       for l in _wrap(linea, ANCHO_CANT, 'Helvetica', 9.5)]
        lineas_desc = _wrap(item['descripcion'], ANCHO_DESC, 'Helvetica', 10)
        # Cada columna se centra sobre su propia cantidad de renglones: si la
        # cantidad ocupa dos líneas y el producto una, el producto igual queda
        # al medio de la fila.
        centro = y_top - alto / 2 - 3.5

        c.setFillColor(c_negro)
        c.setFont('Helvetica', 9.5)
        base_cant = centro + 12.5 * (len(lineas_cant) - 1) / 2
        for i, linea in enumerate(lineas_cant):
            c.drawCentredString((COLS[0] + COLS[1]) / 2, base_cant - 12.5 * i, linea)

        c.setFont('Helvetica', 10)
        base_desc = centro + 12.5 * (len(lineas_desc) - 1) / 2
        for i, linea in enumerate(lineas_desc):
            c.drawCentredString((COLS[1] + COLS[2]) / 2, base_desc - 12.5 * i, linea)

        c.drawCentredString((COLS[2] + COLS[3]) / 2, centro, _gs(item['precio']))
        c.drawCentredString((COLS[3] + COLS[4]) / 2, centro, _gs(item['total']))

    def alto_de(item):
        lineas_cant = [l for linea in _lineas_cantidad(item)
                       for l in _wrap(linea, ANCHO_CANT, 'Helvetica', 9.5)]
        lineas_desc = _wrap(item['descripcion'], ANCHO_DESC, 'Helvetica', 10)
        return max(ALTO_FILA, 16.0 + 12.5 * max(len(lineas_cant), len(lineas_desc)))

    # ── Recorrido de los ítems, con salto de hoja ─────────────────────
    franja_pie()
    encabezado(continuacion=False)
    y = encabezado_tabla()

    pendientes = list(datos['items'])
    while pendientes:
        item = pendientes[0]
        alto = alto_de(item)
        # En la última hoja el piso lo pone la caja de total; en las
        # intermedias se puede bajar hasta la franja del pie.
        piso = TABLA_MIN_Y if len(pendientes) == 1 else FOOTER_H + 14
        if y - alto < piso:
            cerrar_tabla(y)
            c.showPage()
            franja_pie()
            encabezado(continuacion=True)
            y = encabezado_tabla()
            continue
        dibujar_fila(item, y, alto)
        y -= alto
        c.setStrokeColor(c_linea)
        c.setLineWidth(0.5)
        c.line(COLS[0], y, COLS[-1], y)
        pendientes.pop(0)

    # Renglones en blanco hasta la caja de total: la nota se completa a
    # mano cuando se agrega algo en el mostrador.
    while y - ALTO_FILA >= TABLA_MIN_Y:
        y -= ALTO_FILA
        c.setStrokeColor(c_linea)
        c.setLineWidth(0.5)
        c.line(COLS[0], y, COLS[-1], y)

    cerrar_tabla(y)

    # ── Caja de total, pegada al pie de la tabla ──────────────────────
    y_total = y - TOTAL_H
    c.setFillColor(c_beige)
    c.rect(COLS[2], y_total, COLS[-1] - COLS[2], TOTAL_H, stroke=0, fill=1)
    c.setFillColor(c_marron)
    c.setFont('Helvetica-Bold', 11)
    c.drawString(COLS[2] + 12, y_total + 9.5, 'Total:')
    # El monto se achica si hace falta: un total de ocho cifras no entra a
    # 12,5 pt y se montaría sobre la palabra "Total:".
    from reportlab.pdfbase.pdfmetrics import stringWidth
    monto = f"GS. {_gs(datos['total'])}"
    disponible = (COLS[-1] - 12) - (COLS[2] + 12 + stringWidth('Total:', 'Helvetica-Bold', 11) + 8)
    tamanio = 12.5
    while tamanio > 8 and stringWidth(monto, 'Helvetica-Bold', tamanio) > disponible:
        tamanio -= 0.5
    c.setFont('Helvetica-Bold', tamanio)
    c.drawRightString(COLS[-1] - 12, y_total + 9, monto)

    if datos.get('descuento'):
        c.setFillColor(c_gris)
        c.setFont('Helvetica', 9)
        c.drawRightString(COLS[2] - 8, y_total + 9.5,
                          f"Descuento aplicado: Gs. {_gs(datos['descuento'])}")

    # ── Observaciones y vendedor, en la banda libre bajo la tabla ─────
    y_nota = y_total - 16
    if datos.get('observaciones'):
        c.setFillColor(c_marron)
        c.setFont('Helvetica-Bold', 8.5)
        c.drawString(IZQ, y_nota, 'OBSERVACIONES:')
        c.setFillColor(c_gris)
        c.setFont('Helvetica', 9)
        y_nota -= 12
        for linea in _wrap(datos['observaciones'], 320, 'Helvetica', 9)[:2]:
            c.drawString(IZQ, y_nota, linea)
            y_nota -= 11
        y_nota -= 4
    if datos.get('vendedor'):
        c.setFillColor(c_gris)
        c.setFont('Helvetica', 8.5)
        c.drawString(IZQ, max(y_nota, FOOTER_H + 8), f"Atendido por: {datos['vendedor']}")

    c.showPage()
    c.save()
    return buffer.getvalue()



# ══════════════════════════════════════════════════════════════════════
# Excel
# ══════════════════════════════════════════════════════════════════════
FILAS_EXTRA = 12   # renglones en blanco al final, ya con fórmula


def render_xlsx(datos: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Nota de Pedido'
    empresa = _empresa()

    hex_ = lambda color: color.lstrip('#')
    f_beige  = PatternFill('solid', fgColor=hex_(BEIGE))
    borde    = Border(*[Side(style='thin', color='BFBFBF')] * 4)
    borde_gr = Border(*[Side(style='medium', color='000000')] * 4)
    fuente_marron = Font(color=hex_(MARRON), bold=True, size=10)

    # ── Encabezado ────────────────────────────────────────────────────
    for fila, alto in ((1, 22), (2, 22), (3, 20), (4, 20), (5, 16)):
        ws.row_dimensions[fila].height = alto

    if os.path.exists(LOGO):
        img = XLImage(LOGO)
        img.width, img.height = 150, int(150 * 703 / 900)
        ws.add_image(img, 'A1')

    ws.merge_cells('C2:E2')
    ws['C2'] = datos['titulo']
    ws['C2'].font = Font(color=hex_(GRIS), bold=True, size=18)
    ws['C2'].alignment = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('C3:E3')
    ws['C3'] = f"N° {datos.get('numero', '')}"
    ws['C3'].font = Font(color=hex_(MARRON), size=10)
    ws['C3'].alignment = Alignment(horizontal='right')

    ws['D4'] = 'FECHA'
    ws['D4'].font = fuente_marron
    ws['D4'].fill = f_beige
    ws['D4'].alignment = Alignment(horizontal='center')
    ws['D4'].border = borde
    ws['E4'] = datos.get('fecha') or date.today()
    ws['E4'].number_format = 'DD/MM/YYYY'
    ws['E4'].alignment = Alignment(horizontal='center')
    ws['E4'].border = borde

    # ── Datos del cliente (editables) ─────────────────────────────────
    campos = [
        (6, 'CLIENTE:',   datos.get('cliente'),  'TELÉFONO:', datos.get('telefono')),
        (7, 'DIRECCIÓN:', datos.get('direccion'), 'RUC / CI:', datos.get('ruc')),
        (8, 'CIUDAD:',    datos.get('ciudad'),   'VENDEDOR:', datos.get('vendedor')),
    ]
    for fila, lbl_izq, val_izq, lbl_der, val_der in campos:
        ws.cell(row=fila, column=1, value=lbl_izq).font = fuente_marron
        c_val = ws.cell(row=fila, column=2, value=val_izq or '')
        c_val.border = Border(bottom=Side(style='thin', color='BFBFBF'))
        ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=3)
        ws.cell(row=fila, column=4, value=lbl_der).font = fuente_marron
        c_der = ws.cell(row=fila, column=5, value=val_der or '')
        c_der.border = Border(bottom=Side(style='thin', color='BFBFBF'))

    # ── Tabla ─────────────────────────────────────────────────────────
    FILA_TH = 10
    encabezados = ['CANTIDAD', 'UNIDAD', 'PRODUCTO', 'PRECIO UNIT.', 'TOTAL']
    for col, titulo in enumerate(encabezados, 1):
        celda = ws.cell(row=FILA_TH, column=col, value=titulo)
        celda.fill = f_beige
        celda.font = Font(bold=True, size=10.5, color='000000')
        celda.alignment = Alignment(horizontal='center', vertical='center')
        celda.border = borde_gr
    ws.row_dimensions[FILA_TH].height = 22

    fila = FILA_TH + 1
    primera_datos = fila
    for item in datos['items']:
        ws.cell(row=fila, column=1, value=float(item['cantidad']))
        ws.cell(row=fila, column=2, value=item.get('unidad') or '')
        ws.cell(row=fila, column=3, value=item['descripcion'])
        ws.cell(row=fila, column=4, value=float(item['precio']))
        fila += 1

    # Renglones en blanco: la nota se sigue completando a mano en Excel.
    ultima_datos = fila + FILAS_EXTRA - 1

    for f in range(primera_datos, ultima_datos + 1):
        # TOTAL por fórmula; el IF evita que los renglones vacíos muestren 0.
        ws.cell(row=f, column=5, value=f'=IF(A{f}="","",A{f}*D{f})')
        for col in range(1, 6):
            celda = ws.cell(row=f, column=col)
            celda.border = borde
            celda.font = Font(size=10)
        ws.cell(row=f, column=1).number_format = '#,##0.00'
        ws.cell(row=f, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=f, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=f, column=3).alignment = Alignment(vertical='center', wrap_text=True)
        ws.cell(row=f, column=4).number_format = '#,##0'
        ws.cell(row=f, column=5).number_format = '#,##0'

    # ── Totales ───────────────────────────────────────────────────────
    fila_sub = ultima_datos + 1
    suma = f'=SUM(E{primera_datos}:E{ultima_datos})'
    if datos.get('descuento'):
        filas_total = [
            ('SUBTOTAL',  suma, False),
            ('DESCUENTO', -float(datos['descuento']), False),
            ('TOTAL',     f'=E{fila_sub}+E{fila_sub + 1}', True),
        ]
    else:
        filas_total = [('TOTAL', suma, True)]

    for i, (etiqueta, valor, destacado) in enumerate(filas_total):
        f = fila_sub + i
        ws.merge_cells(start_row=f, start_column=3, end_row=f, end_column=4)
        c_lbl = ws.cell(row=f, column=3, value=etiqueta)
        c_lbl.alignment = Alignment(horizontal='right', vertical='center')
        c_val = ws.cell(row=f, column=5, value=valor)
        c_val.number_format = '#,##0'
        c_val.alignment = Alignment(horizontal='right', vertical='center')
        if destacado:
            for col in (3, 4, 5):
                ws.cell(row=f, column=col).fill = f_beige
                ws.cell(row=f, column=col).border = borde_gr
            c_lbl.font = Font(bold=True, size=11, color=hex_(MARRON))
            c_val.font = Font(bold=True, size=12, color=hex_(MARRON))
            c_val.number_format = '"Gs. "#,##0'
        else:
            c_lbl.font = Font(bold=True, size=10, color=hex_(GRIS))
            c_val.font = Font(size=10)

    fila = fila_sub + len(filas_total) + 2

    if datos.get('observaciones'):
        ws.cell(row=fila, column=1, value='OBSERVACIONES:').font = fuente_marron
        ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=5)
        c_obs = ws.cell(row=fila, column=2, value=datos['observaciones'])
        c_obs.alignment = Alignment(wrap_text=True, vertical='top')
        fila += 2

    # ── Pie de contacto ───────────────────────────────────────────────
    ws.cell(row=fila, column=1, value='CONTACTO:').font = Font(bold=True, size=10, color=hex_(MARRON))
    for col in range(1, 6):
        ws.cell(row=fila, column=col).fill = f_beige
    for texto in (empresa['email'], empresa['direccion'], empresa['telefono']):
        fila += 1
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
        celda = ws.cell(row=fila, column=1, value=texto)
        celda.font = Font(size=9, color=hex_(MARRON))
        for col in range(1, 6):
            ws.cell(row=fila, column=col).fill = f_beige

    # ── Presentación ──────────────────────────────────────────────────
    for col, ancho in zip('ABCDE', (12, 18, 52, 15, 18)):
        ws.column_dimensions[col].width = ancho
    ws.freeze_panes = f'A{FILA_TH + 1}'
    ws.print_area = f'A1:{get_column_letter(5)}{fila}'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    salida = io.BytesIO()
    wb.save(salida)
    return salida.getvalue()


# ══════════════════════════════════════════════════════════════════════
# Entrega HTTP
# ══════════════════════════════════════════════════════════════════════
def responder(datos: dict, formato: str) -> HttpResponse:
    referencia = datos.get('numero') or datetime.now().strftime('%Y%m%d%H%M')
    nombre = f"nota_{datos.get('tipo', TIPO_DEFECTO)}_{referencia}"
    if formato == 'xlsx':
        contenido = render_xlsx(datos)
        mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ext = 'xlsx'
    else:
        contenido = render_pdf(datos)
        mime = 'application/pdf'
        ext = 'pdf'
    resp = HttpResponse(contenido, content_type=mime)
    resp['Content-Disposition'] = f'attachment; filename="{nombre}.{ext}"'
    return resp
