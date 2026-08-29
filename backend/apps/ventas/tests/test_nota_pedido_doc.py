"""
Tests de la Nota de Pedido como documento (PDF / Excel).

No comparan píxeles: verifican que el documento se genere, que arrastre los
datos del pedido y que el Excel salga con fórmulas (que es lo que lo hace
editable, el motivo de que exista ese formato).
"""
from decimal import Decimal

from django.test import TestCase
from openpyxl import load_workbook

from apps.inventario.models import MovimientoStock, Stock
from apps.productos.models import Producto
from apps.productos.tests.factories import crear_producto, crear_usuario, crear_variante
from apps.ventas import nota_pedido_doc as doc
from apps.ventas.models import ItemPedido, NotaPedido


class NotaPedidoDocumentoTests(TestCase):
    # Crear catalogo dispara el registro de cambios del sync (base aparte).
    databases = {'default', 'sync'}

    def setUp(self):
        self.usuario = crear_usuario('vendedora')
        producto = crear_producto(nombre='Porcelanato Ébano')
        self.variante = crear_variante(
            producto=producto, color='Ébano',
            largo_cm=Decimal('120'), ancho_cm=Decimal('60'), piezas_por_caja=2,
        )
        stock = Stock.objects.get(variante=self.variante)
        stock.registrar_movimiento(
            tipo=MovimientoStock.TIPO_ENTRADA, cantidad=Decimal('100'),
            usuario=self.usuario, observaciones='carga de prueba',
        )

        self.pedido = NotaPedido.objects.create(
            vendedor=self.usuario, cliente_nombre='Maru López',
            cliente_telefono='0971 000000',
        )
        ItemPedido.objects.create(
            pedido=self.pedido, variante=self.variante,
            cantidad=Decimal('24.99'), precio_unitario=Decimal('97500'),
        )
        self.pedido.recalcular_totales()

    # ── Armado de los datos ──────────────────────────────────────────
    def test_titulo_por_defecto_es_presupuesto(self):
        """Es el documento que se le pasa al cliente antes de cerrar la venta."""
        self.assertEqual(doc.datos_desde_pedido(self.pedido)['titulo'],
                         'NOTA DE PRESUPUESTO')

    def test_titulo_de_pedido_cuando_se_pide_ese_tipo(self):
        datos = doc.datos_desde_pedido(self.pedido, tipo=doc.TIPO_PEDIDO)
        self.assertEqual(datos['titulo'], 'NOTA DE PEDIDO')

    def test_tipo_desconocido_cae_al_de_defecto(self):
        datos = doc.datos_desde_pedido(self.pedido, tipo='cualquier-cosa')
        self.assertEqual(datos['tipo'], doc.TIPO_PRESUPUESTO)

    def test_datos_arrastran_cliente_y_totales(self):
        datos = doc.datos_desde_pedido(self.pedido)

        self.assertEqual(datos['cliente'], 'Maru López')
        self.assertEqual(datos['numero'], self.pedido.numero)
        self.assertEqual(len(datos['items']), 1)
        self.assertEqual(datos['total'], Decimal('2436525.00'))

    def test_unidad_de_m2_indica_cuantas_cajas(self):
        """24,99 m² con cajas de 1,44 m² son 18 cajas (se redondea para arriba)."""
        datos = doc.datos_desde_pedido(self.pedido)
        self.assertEqual(datos['items'][0]['unidad'], 'mts2 (18 cajas)')

    def test_producto_por_pieza_no_lleva_unidad(self):
        self.variante.producto.unidad_venta = Producto.UNIDAD_PIEZA
        self.variante.producto.save(update_fields=['unidad_venta'])

        datos = doc.datos_desde_pedido(self.pedido)
        self.assertEqual(datos['items'][0]['unidad'], '')

    # ── Render ───────────────────────────────────────────────────────
    def test_pdf_se_genera(self):
        contenido = doc.render_pdf(doc.datos_desde_pedido(self.pedido))

        self.assertTrue(contenido.startswith(b'%PDF'))
        self.assertGreater(len(contenido), 5000)

    def test_excel_deja_el_total_como_formula(self):
        """
        El Excel es el documento de trabajo: si el total viniera calculado y
        fijo, editar una cantidad no recalcularía nada.
        """
        from io import BytesIO
        wb = load_workbook(BytesIO(doc.render_xlsx(doc.datos_desde_pedido(self.pedido))))
        ws = wb.active

        primera = 11   # la fila siguiente al encabezado de la tabla
        self.assertEqual(ws.cell(row=primera, column=1).value, 24.99)
        self.assertEqual(ws.cell(row=primera, column=4).value, 97500)
        self.assertTrue(str(ws.cell(row=primera, column=5).value).startswith('=IF('))

        formulas = [c.value for col in ws.iter_cols(min_col=5, max_col=5) for c in col]
        self.assertTrue(any(str(v).startswith('=SUM(') for v in formulas))

    def test_excel_deja_renglones_en_blanco_para_seguir_cargando(self):
        from io import BytesIO
        wb = load_workbook(BytesIO(doc.render_xlsx(doc.datos_desde_pedido(self.pedido))))
        ws = wb.active

        vacias = [f for f in range(12, 12 + doc.FILAS_EXTRA)
                  if ws.cell(row=f, column=1).value is None
                  and str(ws.cell(row=f, column=5).value).startswith('=IF(')]
        self.assertEqual(len(vacias), doc.FILAS_EXTRA)


class NotaPedidoEndpointTests(TestCase):
    """El cableado de la URL y los permisos, que es lo que más fácil se rompe."""

    databases = {'default', 'sync'}

    def setUp(self):
        from rest_framework.test import APIClient

        self.usuario = crear_usuario('vendedora', rol='vendedor')
        self.deposito = crear_usuario('depositero', rol='deposito')
        producto = crear_producto(nombre='Porcelanato Ébano')
        variante = crear_variante(producto=producto, color='Ébano')
        Stock.objects.get(variante=variante).registrar_movimiento(
            tipo=MovimientoStock.TIPO_ENTRADA, cantidad=Decimal('50'),
            usuario=self.usuario, observaciones='carga de prueba',
        )
        self.pedido = NotaPedido.objects.create(
            vendedor=self.usuario, cliente_nombre='Maru López',
        )
        ItemPedido.objects.create(
            pedido=self.pedido, variante=variante,
            cantidad=Decimal('3'), precio_unitario=Decimal('165000'),
        )
        self.pedido.recalcular_totales()
        self.client = APIClient()

    def _url(self, formato):
        return f'/api/v1/ventas/pedidos/{self.pedido.id}/nota/?formato={formato}'

    def test_vendedor_descarga_el_pdf(self):
        self.client.force_authenticate(self.usuario)
        resp = self.client.get(self._url('pdf'))

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')
        self.assertIn('nota_presupuesto_', resp['Content-Disposition'])

    def test_el_tipo_llega_al_nombre_del_archivo(self):
        self.client.force_authenticate(self.usuario)
        resp = self.client.get(self._url('pdf') + '&tipo=pedido')

        self.assertEqual(resp.status_code, 200)
        self.assertIn('nota_pedido_', resp['Content-Disposition'])

    def test_vendedor_descarga_el_excel(self):
        self.client.force_authenticate(self.usuario)
        resp = self.client.get(self._url('xlsx'))

        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml', resp['Content-Type'])

    def test_deposito_no_accede_porque_el_documento_lleva_precios(self):
        self.client.force_authenticate(self.deposito)
        self.assertEqual(self.client.get(self._url('pdf')).status_code, 403)
