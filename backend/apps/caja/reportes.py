"""
apps/caja/reportes.py
Generación de reportes en PDF (reportlab) y Excel (openpyxl).
Tres reportes: Stock, Balance de Ventas, Extracto de Caja.

Cada función de reporte arma una estructura común:
  { 'titulo', 'subtitulo', 'columnas': [...], 'filas': [[...]], 'totales': {...} }
y luego se renderiza a PDF o XLSX con los helpers de abajo.
"""
import io
from datetime import datetime, date
from django.http import HttpResponse
from django.db.models import Sum


# ════════════════════════════════════════════════════════
# Helpers de formato
# ════════════════════════════════════════════════════════
def _gs(v):
    try:
        return f'Gs. {int(v):,}'.replace(',', '.')
    except (ValueError, TypeError):
        return 'Gs. 0'


# ════════════════════════════════════════════════════════
# Render a PDF (reportlab)
# ════════════════════════════════════════════════════════
def render_pdf(reporte: dict, tamanio: str = 'a4') -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm, inch
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )
    from xml.sax.saxutils import escape

    # "Oficio" — tamaño de hoja de uso administrativo habitual en Paraguay
    # (8.5 × 13 pulgadas). Distinto del A4 (210 × 297 mm) y no incluido en
    # reportlab.lib.pagesizes (que solo trae LEGAL = 8.5 × 14").
    OFICIO = (8.5 * inch, 13 * inch)

    GOLD     = colors.HexColor('#B99C74')
    GOLDDARK = colors.HexColor('#8a7355')
    SIDEBAR  = colors.HexColor('#453941')
    BORDER   = colors.HexColor('#e8e4df')
    TEXTSEC  = colors.HexColor('#6b6560')
    TEXT     = colors.HexColor('#1a1714')

    buffer = io.BytesIO()
    muchas_cols = len(reporte['columnas']) > 5
    base_pagesize = OFICIO if tamanio == 'oficio' else A4
    pagesize = landscape(base_pagesize) if muchas_cols else base_pagesize
    doc = SimpleDocTemplate(buffer, pagesize=pagesize,
                            topMargin=18*mm, bottomMargin=20*mm,
                            leftMargin=14*mm, rightMargin=14*mm)

    styles = getSampleStyleSheet()
    st_titulo = ParagraphStyle('titulo', parent=styles['Title'],
        fontSize=18, textColor=SIDEBAR, spaceAfter=2, alignment=0)
    st_marca = ParagraphStyle('marca', parent=styles['Normal'],
        fontSize=10, textColor=GOLDDARK, spaceAfter=1)
    st_sub = ParagraphStyle('sub', parent=styles['Normal'],
        fontSize=9, textColor=TEXTSEC, spaceAfter=12)
    st_th = ParagraphStyle('th', parent=styles['Normal'],
        fontSize=8.5, textColor=GOLD, fontName='Helvetica-Bold', leading=11)
    st_celda = ParagraphStyle('celda', parent=styles['Normal'],
        fontSize=8, textColor=TEXT, leading=10)
    st_celda_der = ParagraphStyle('celda_der', parent=st_celda, alignment=2)

    elementos = []
    elementos.append(Paragraph('ÓGA PORÃ — Acabados de Construcción', st_marca))
    elementos.append(Paragraph(reporte['titulo'], st_titulo))
    sub = reporte.get('subtitulo', '')
    generado = datetime.now().strftime('%d/%m/%Y %H:%M')
    elementos.append(Paragraph(f'{sub}  ·  Generado: {generado}', st_sub))

    # ── Tabla ────────────────────────────────────────────────
    # colWidths explícitos: sin esto reportlab no envuelve el texto y da
    # a cada columna su ancho "natural", desbordando la hoja apenas hay
    # nombres largos de producto o cliente. Las celdas van en Paragraph
    # (en vez de strings planos) para que sí hagan wrap dentro del ancho
    # asignado, y escapadas porque son texto libre cargado por el usuario.
    columnas    = reporte['columnas']
    cols_der    = set(reporte.get('cols_derecha', []))
    n_cols      = len(columnas)
    ancho_total = pagesize[0] - doc.leftMargin - doc.rightMargin

    ancho_num   = ancho_total * 0.09
    n_texto     = max(n_cols - len(cols_der), 1)
    ancho_texto = (ancho_total - ancho_num * len(cols_der)) / n_texto
    col_widths  = [ancho_num if i in cols_der else ancho_texto for i in range(n_cols)]

    header = [Paragraph(escape(str(c)), st_th) for c in columnas]
    filas_pdf = []
    for fila in reporte['filas']:
        fila_pdf = []
        for i, valor in enumerate(fila):
            estilo_celda = st_celda_der if i in cols_der else st_celda
            fila_pdf.append(Paragraph(escape(str(valor)), estilo_celda))
        filas_pdf.append(fila_pdf)

    tabla = Table([header] + filas_pdf, repeatRows=1, colWidths=col_widths)
    tabla.setStyle(TableStyle([
        ('BACKGROUND',     (0,0), (-1,0), SIDEBAR),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#fafaf9')]),
        ('GRID',           (0,0), (-1,-1), 0.5, BORDER),
        ('VALIGN',         (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',     (0,0), (-1,-1), 5),
        ('BOTTOMPADDING',  (0,0), (-1,-1), 5),
        ('LEFTPADDING',    (0,0), (-1,-1), 6),
        ('RIGHTPADDING',   (0,0), (-1,-1), 6),
    ]))
    elementos.append(tabla)

    # Totales
    if reporte.get('totales'):
        elementos.append(Spacer(1, 10))
        for k, v in reporte['totales'].items():
            elementos.append(Paragraph(
                f'<b>{escape(str(k))}:</b> {escape(str(v))}',
                ParagraphStyle('tot', parent=styles['Normal'],
                    fontSize=10, textColor=SIDEBAR, alignment=2, spaceAfter=2)))

    def _pie_pagina(canvas, doc_):
        """Pie con marca y número de página — evita hojas sueltas sin identificar."""
        canvas.saveState()
        canvas.setFont('Helvetica', 7.5)
        canvas.setFillColor(TEXTSEC)
        ancho_pagina = doc_.pagesize[0]
        canvas.drawString(doc_.leftMargin, 10*mm, 'Oga Porã — Documento de uso interno')
        canvas.drawRightString(ancho_pagina - doc_.rightMargin, 10*mm, f'Página {doc_.page}')
        canvas.restoreState()

    doc.build(elementos, onFirstPage=_pie_pagina, onLaterPages=_pie_pagina)
    return buffer.getvalue()


# ════════════════════════════════════════════════════════
# Render a Excel (openpyxl)
# ════════════════════════════════════════════════════════
def render_xlsx(reporte: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = reporte.get('hoja', 'Reporte')[:31]

    GOLD_FILL = PatternFill('solid', fgColor='453941')
    GOLD_FONT = Font(color='B99C74', bold=True, size=11)
    borde = Border(*[Side(style='thin', color='E8E4DF')]*4)

    # Encabezado del documento
    ws['A1'] = 'ÓGA PORÃ — Acabados de Construcción'
    ws['A1'].font = Font(color='8a7355', bold=True, size=12)
    ws['A2'] = reporte['titulo']
    ws['A2'].font = Font(color='453941', bold=True, size=14)
    ws['A3'] = f"{reporte.get('subtitulo','')}  ·  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A3'].font = Font(color='6b6560', size=9)

    fila_header = 5
    # Cabecera de columnas
    for col_idx, titulo in enumerate(reporte['columnas'], 1):
        c = ws.cell(row=fila_header, column=col_idx, value=titulo)
        c.fill = GOLD_FILL; c.font = GOLD_FONT
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = borde

    # Filas de datos
    for r_idx, fila in enumerate(reporte['filas'], fila_header+1):
        for c_idx, valor in enumerate(fila, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=valor)
            c.border = borde
            c.font = Font(size=10)
            if (c_idx-1) in reporte.get('cols_derecha', []):
                c.alignment = Alignment(horizontal='right')

    # Anchos de columna automáticos
    for col_idx, titulo in enumerate(reporte['columnas'], 1):
        max_len = len(str(titulo))
        for fila in reporte['filas']:
            if col_idx-1 < len(fila):
                max_len = max(max_len, len(str(fila[col_idx-1])))
        ws.column_dimensions[ws.cell(row=fila_header, column=col_idx).column_letter].width = min(max_len+3, 45)

    # Totales
    if reporte.get('totales'):
        fila_tot = fila_header + len(reporte['filas']) + 2
        for k, v in reporte['totales'].items():
            ws.cell(row=fila_tot, column=1, value=k).font = Font(bold=True, color='453941')
            ws.cell(row=fila_tot, column=2, value=v).font = Font(bold=True)
            fila_tot += 1

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ════════════════════════════════════════════════════════
# Construcción de cada reporte
# ════════════════════════════════════════════════════════
def reporte_stock():
    from apps.inventario.models import Stock
    qs = Stock.objects.select_related('variante__producto').all()

    filas = []
    total_unidades = 0
    sin_stock = 0
    criticos = 0
    for s in qs:
        disp = float(s.cantidad_disponible)
        total_unidades += disp
        if s.sin_stock: sin_stock += 1
        elif s.en_stock_critico: criticos += 1
        estado = 'Sin stock' if s.sin_stock else ('Crítico' if s.en_stock_critico else 'Disponible')
        filas.append([
            s.variante.producto.nombre,
            s.variante.sku,
            str(s.variante),
            f'{float(s.cantidad):.2f}',
            f'{float(s.cantidad_reservada):.2f}',
            f'{disp:.2f}',
            f'{float(s.stock_minimo):.2f}',
            estado,
        ])
    filas.sort(key=lambda f: f[0])

    return {
        'titulo':    'Reporte de Stock',
        'subtitulo': f'{qs.count()} variantes registradas',
        'hoja':      'Stock',
        'columnas':  ['Producto','SKU','Variante','Stock','Reservado','Disponible','Mínimo','Estado'],
        'cols_derecha': [3,4,5,6],
        'filas':     filas,
        'totales': {
            'Variantes sin stock':  sin_stock,
            'Variantes en crítico': criticos,
            'Total unidades disponibles': f'{total_unidades:.2f}',
        },
    }


def reporte_ventas(desde: date, hasta: date):
    from apps.caja.models import Pago

    pagos = Pago.objects.filter(
        estado=Pago.ESTADO_CONFIRMADO,
        fecha__date__gte=desde,
        fecha__date__lte=hasta,
    ).select_related('pedido', 'cajero').order_by('fecha')

    filas = []
    total_general = 0
    por_medio = {}
    for p in pagos:
        monto = float(p.monto)
        total_general += monto
        medio = p.get_medio_pago_display()
        por_medio[medio] = por_medio.get(medio, 0) + monto
        filas.append([
            p.fecha.strftime('%d/%m/%Y %H:%M'),
            p.numero_ticket or '—',
            p.pedido.numero if p.pedido else '—',
            p.pedido.cliente_nombre if p.pedido and p.pedido.cliente_nombre else 'Consumidor Final',
            medio,
            p.cajero.nombre_completo,
            _gs(monto),
        ])

    totales = {'Total de ventas': _gs(total_general),
               'Cantidad de cobros': len(filas)}
    for medio, m in por_medio.items():
        totales[f'  · {medio}'] = _gs(m)

    return {
        'titulo':    'Balance de Ventas',
        'subtitulo': f'Del {desde.strftime("%d/%m/%Y")} al {hasta.strftime("%d/%m/%Y")}',
        'hoja':      'Ventas',
        'columnas':  ['Fecha','Ticket','Pedido','Cliente','Medio','Cajero','Monto'],
        'cols_derecha': [6],
        'filas':     filas,
        'totales':   totales,
    }


def reporte_caja(desde: date, hasta: date):
    from apps.caja.models import SesionCaja, Pago

    sesiones = SesionCaja.objects.filter(
        fecha_apertura__date__gte=desde,
        fecha_apertura__date__lte=hasta,
    ).select_related('cajero').order_by('fecha_apertura')

    filas = []
    total_ventas = 0
    for s in sesiones:
        ventas = Pago.objects.filter(
            sesion_caja=s, estado=Pago.ESTADO_CONFIRMADO
        ).aggregate(t=Sum('monto'))['t'] or 0
        total_ventas += float(ventas)
        filas.append([
            s.fecha_apertura.strftime('%d/%m/%Y %H:%M'),
            s.fecha_cierre.strftime('%d/%m/%Y %H:%M') if s.fecha_cierre else 'Abierta',
            s.cajero.nombre_completo,
            _gs(s.monto_apertura),
            _gs(s.monto_cierre) if s.monto_cierre is not None else '—',
            _gs(ventas),
            'Cerrada' if s.estado == 'cerrada' else 'Abierta',
        ])

    return {
        'titulo':    'Extracto de Caja',
        'subtitulo': f'Del {desde.strftime("%d/%m/%Y")} al {hasta.strftime("%d/%m/%Y")}',
        'hoja':      'Caja',
        'columnas':  ['Apertura','Cierre','Cajero','M. Apertura','M. Cierre','Ventas','Estado'],
        'cols_derecha': [3,4,5],
        'filas':     filas,
        'totales': {
            'Sesiones': len(filas),
            'Total ventas del período': _gs(total_ventas),
        },
    }


# ════════════════════════════════════════════════════════
# Entrega como HttpResponse
# ════════════════════════════════════════════════════════
def responder_reporte(reporte: dict, formato: str, nombre_base: str, tamanio: str = 'a4') -> HttpResponse:
    hoy = date.today().strftime('%Y%m%d')
    if formato == 'pdf':
        contenido = render_pdf(reporte, tamanio=tamanio)
        resp = HttpResponse(contenido, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{nombre_base}_{hoy}.pdf"'
        return resp
    else:  # xlsx
        contenido = render_xlsx(reporte)
        resp = HttpResponse(
            contenido,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{nombre_base}_{hoy}.xlsx"'
        return resp
